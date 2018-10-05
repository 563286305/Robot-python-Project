# encoding: utf-8
__author__ = 'zhengyong'

import os
import sys
from openpyxl import Workbook, load_workbook
from contextlib import closing
import win32com.client as win32
from openpyxl.styles import Font, Color, Alignment
from openpyxl.styles import colors
from openpyxl import Workbook
import csv
import codecs
import re


reload(sys)
sys.setdefaultencoding('utf-8')


class ReadWriteExcel():

	# 创建Excel文档和sheet
	def create_excel(self, file_name, sheet_name):
		with closing(Workbook()) as wb:
			ws = wb.worksheets[0]
			ws.title = str(sheet_name)
			wb.save(file_name)

	# 拷贝Excel中sheet内容
	def copysheet(self,fromfile_name, tofile_name, fromsheet_name, tosheet_name):
		wb1 = load_workbook(str(fromfile_name))
		wb2 = load_workbook(str(tofile_name))
		ft = Font(name="微软雅黑", size=10)
		sheets = wb2.sheetnames
		if tosheet_name in sheets:
			wb2[str(tosheet_name)]
		else:
			wb2.create_sheet(str(tosheet_name))
		ws1 = wb1[str(fromsheet_name)]
		ws2 = wb2[str(tosheet_name)]

		for i,row in enumerate(ws1.iter_rows()):
			for j,cell in enumerate(row):
				ws2.cell(row=i + 1, column=j + 1, value=cell.value).font = ft
		wb2.save(tofile_name)

	#根据单元格内容获取单元格行列位置信息（按行），用来获取产品所在的列名
	def get_cell_index(self,file_name, sheet_name, rowNo):
		with closing(load_workbook(filename=file_name)) as wb:
			sheets = wb.sheetnames
			if sheet_name in sheets:
				ws = wb[str(sheet_name)]
			else:
				ws = wb.create_sheet(str(sheet_name))
			num = int(rowNo) - 1
			tagdict = {}
			for cell in list(ws.rows)[num]:
				if cell.value != None:
					value = cell.value
					value = str(value)
#					print "value:    "+ value
					index = cell.coordinate
#					print "index: " + index
					indexlist = re.findall('[A-Z]', index)
					colindex = ''.join(indexlist)
				tagdict[value] = colindex
		return 	tagdict

	# 获取指定sheet中行数
	def get_excel_row_count(self, file_name, sheet_name):
		with closing(load_workbook(filename=file_name)) as wb:
#			rows = wb.get_sheet_by_name(name=str(sheet_name)).max_row
			rows = wb[str(sheet_name)].max_row
			return rows

	# 获取指定sheet中非空行数
	def get_excel_no_blank_row_count(self, file_name, sheet_name, row_count):
		with closing(load_workbook(filename=file_name)) as wb:
#			ws = wb.get_sheet_by_name(str(sheet_name))
			ws = wb[str(sheet_name)]

			colValueList = []
			num = int(row_count)
			for cell in list(ws.columns)[num]:
				if cell.value != None:
					values = cell.value
					colValueList.append(values)
				else:
					pass
			return len(colValueList)

	# 获取指定sheet中列数
	def get_excel_column_count(self, file_name, sheet_name):
		with closing(load_workbook(filename=file_name)) as wb:
			columns = wb.get_sheet_by_name(name=str(sheet_name)).max_column
			return columns
	#读取两个指定列，并组成字典
	def get_excel_name_value_dict(self, file_name, sheet_name,colname1,colname2,row_start,row_end):

		namelist = self.colname_read_excel_return_list(file_name,sheet_name,colname1,row_start, row_end)
		valuelist = self.colname_read_excel_return_list(file_name,sheet_name,colname2,row_start, row_end)

		return dict(zip(namelist, valuelist))

	def readExcel(self, file_name, sheet_name, cell_row, cell_column):
		with closing(load_workbook(filename=file_name, data_only=True)) as wb:
			ws = wb.active
			ws.title = str(sheet_name)
			cellValue = str(ws.cell(row=int(cell_row), column=int(cell_column)).value)
			return cellValue.decode('utf-8')

	def addExcelwithFormat(self, file_name, sheet_name, cell_row, cell_column, value):
		with closing(load_workbook(filename=file_name)) as wb:
			ws = wb[str(sheet_name)]
			ft1 = Font(name="微软雅黑", color=colors.RED, size=10)
			ft2 = Font(name="微软雅黑", size=10, bold=True)
			ft3 = Font(name="微软雅黑 Light", size=9)
			if value == 'FAIL':
				ws.cell(row=int(cell_row), column=int(cell_column), value=value).font = ft1
			elif value == 'PASS':
				ws.cell(row=int(cell_row), column=int(cell_column), value=value).font = ft2
			else:
				ws.cell(row=int(cell_row), column=int(cell_column), value=value).font = ft3
			wb.save(file_name)


	def addExcel(self, file_name, sheet_name, cell_row, cell_column, value):
		with closing(load_workbook(filename=file_name)) as wb:
#			ws = wb.get_sheet_by_name(str(sheet_name))
			ws = wb[str(sheet_name)]
			ft = Font(name="微软雅黑 Light", size=9)
			ws.cell(row=int(cell_row), column=int(cell_column), value=value).font = ft
			wb.save(file_name)


	def addTitleExcel(self, file_name, sheet_name, cell_row, cell_column, value):
		with closing(load_workbook(filename=file_name)) as wb:
#			ws = wb.get_sheet_by_name(str(sheet_name))
			ws = wb[str(sheet_name)]
			ft = Font(name="微软雅黑", size=10, bold=True)
			ws.cell(row=int(cell_row), column=int(cell_column), value=value).font = ft
			wb.save(file_name)


	def listReadExcel(self, file_name, sheet_name, cell_row, cell_column, start_col):
		with closing(load_workbook(filename=file_name, data_only=True)) as wb:
#			ws = wb.get_sheet_by_name(str(sheet_name))
			ws = wb[str(sheet_name)]
			cellValueList = []
			for i in range(cell_column):
				cellValue = ws.cell(row=int(cell_row), column=int(start_col) + i).value
				cellValueList.append(cellValue)
			return cellValueList


	def list_add_excel(self, file_name, sheet_name, cell_row, cell_column, *value):
		with closing(load_workbook(filename=file_name)) as wb:
			sheets = wb.sheetnames
			if sheet_name in sheets:
				ws = wb[str(sheet_name)]
			else:
				ws = wb.create_sheet(str(sheet_name))
			ft1 = Font(name="微软雅黑", color=colors.RED, size=10)
			ft2 = Font(name="微软雅黑", size=10, bold=True)
			ft3 = Font(name="微软雅黑 Light", size=9)
			for i in range(len(value)):
				if value[i] == 'FAIL':
					ws.cell(row=int(cell_row), column=int(cell_column) + i, value=value[i]).font = ft1
				elif value[i] == 'PASS':
					ws.cell(row=int(cell_row), column=int(cell_column) + i, value=value[i]).font = ft2
				else:
					ws.cell(row=int(cell_row), column=int(cell_column) + i, value=value[i]).font = ft3
			wb.save(file_name)

	def list_write_title_excel(self, open_file_name, sheet_name, cell_row, cell_column, save_file_name, *value):
		with closing(load_workbook(filename=open_file_name)) as wb:
			sheets = wb.sheetnames
			if sheet_name in sheets:
				ws = wb[str(sheet_name)]
			else:
				ws = wb.create_sheet(str(sheet_name))
			ft = Font(name="微软雅黑", size=10, bold=True)
			align = Alignment(horizontal='center', vertical='bottom')
			for i in range(len(value)):
				ws.cell(row=int(cell_row), column=int(cell_column) + i, value=value[i]).font = ft
				ws.cell(row=int(cell_row), column=int(cell_column) + i, value=value[i]).alignment = align
			wb.save(save_file_name)

	# 刷新Excel文件
	def refreshExcel(self, file_name):
		excel = win32.gencache.EnsureDispatch('Excel.Application')
		excel.Visible = False
		workbook = excel.Workbooks.Open(os.path.join(os.getcwd(), file_name))
		workbook.Save()
		excel.Application.Quit()


	# 按照单元格名称写入Excel
	def cellAddExcel(self, file_name, sheet_name, cell, value):
		with closing(load_workbook(filename=file_name)) as wb:
			ws = wb[str(sheet_name)]
			ft = Font(name="微软雅黑 Light", size=9, color=colors.GREEN)
			ws[str(cell)] = value
			ws[str(cell)].font = ft
			wb.save(file_name)


	# 通过单元格名称写入Excel
	def cellListAddExcel(self, file_name, sheet_name, value, *cell):
		with closing(load_workbook(filename=file_name)) as wb:
			ws = wb[str(sheet_name)]
			ft = Font(name="微软雅黑 Light", size=10, bold=True, color=colors.GREEN)
			for i in range(len(cell)):
				ws[str(cell[i])] = value
				ws[str(cell[i])].font = ft
			wb.save(file_name)


	# 通过单元格名称读取Excel
	def cellListReadExcel(self, file_name, sheet_name, col_name, start, end):
		with closing(load_workbook(filename=file_name, data_only=True)) as wb:
			ws = wb[str(sheet_name)]
			cellValueList = []
			for i in range(int(start), int(end)):
				cellname = str(col_name) + str(i)
				cellValue = ws[str(cellname)].value
				if cellValue != None:
					cellValueList.append(cellValue)
				else:
					break
			return cellValueList


	# # 拷贝Excel中内容
	# def excelCopy(self, fromfile_name, tofile_name, fromsheet_name, tosheet_name):
	# 	wb1 = load_workbook(str(fromfile_name))
	# 	wb2 = load_workbook(str(tofile_name))
	# 	ws1 = wb1[str(fromsheet_name)]
	# 	ws2 = wb2[str(tosheet_name)]
	# 	ft = Font(name="微软雅黑 Light", size=9)
	# 	for i, row in enumerate(ws1.iter_rows()):
	# 		for j, cell in enumerate(row):
	# 			ws2.cell(row=i + 1, column=j + 1, value=cell.value).font = ft
	# 	wb2.save(tofile_name)


	# 按行读取Excel中内容
	def rowsReadExcel(self, file_name, sheet_name, col_count):
		with closing(load_workbook(filename=file_name, data_only=True)) as wb:
			ws = wb[str(sheet_name)]
			rowValueList = []
			num = int(col_count)
			for cell in list(ws.rows)[num]:
				values = cell.value
				rowValueList.append(values)
			return rowValueList


	# 按行读取Excel中非空内容
	def rowsReadExcel2(self, file_name, sheet_name, col_count):
		with closing(load_workbook(filename=file_name, data_only=True)) as wb:
			ws = wb[str(sheet_name)]
			rowValueList = []
			num = int(col_count)
			for cell in list(ws.rows)[num]:
				if cell.value != None:
					values = cell.value
				else:
					break
				rowValueList.append(values)
			return rowValueList


	# 按列读取Excel中内容
	def cols_read_excel(self, file_name, sheet_name, row_count):
		with closing(load_workbook(filename=file_name, data_only=True)) as wb:
			ws = wb[str(sheet_name)]
			col_value_list = []
			num = int(row_count)
			for cell in list(ws.columns)[num]:
				if cell.value != None:
					values = cell.value
					col_value_list.append(str(values))
				else:
					pass
			return col_value_list

	# 按列读取Excel中所有内容，包括空
	def cols_read_excela_all(self, file_name, sheet_name, row_count):
		with closing(load_workbook(filename=file_name, data_only=True)) as wb:
			ws = wb[str(sheet_name)]
			col_value_list = []
			num = int(row_count)
			for cell in list(ws.columns)[num]:
				values = cell.value
				col_value_list.append(str(values))
			return col_value_list
	#复制excel
	def copyExcel(self,file_name, save_name):
		with closing(load_workbook(filename=file_name)) as wb:
			wb.save(save_name)

	# 将某个产品的测试结果写入到Excel，按 dict={'2': 'PASS', '4': 'PASS'},并保存到新的excel里
	def resultdictAddExcel(self, file_name, sheet_name, resultdict, col):
		with closing(load_workbook(filename=file_name)) as wb:
			sheets = wb.sheetnames
			if sheet_name in sheets:
				ws = wb[str(sheet_name)]
			else:
				ws = wb.create_sheet(str(sheet_name))
			ft1 = Font(name="微软雅黑", color=colors.RED, size=10)
			ft2 = Font(name="微软雅黑", size=10, bold=True)
			ft3 = Font(name="微软雅黑 Light", size=9)
			for key in resultdict.keys():
				cellname = col + str(key)
				ws[str(cellname)] = resultdict[key]
				if resultdict[key] == 'PASS':
					ws[str(cellname)].font = ft2
				if resultdict[key] == 'FAILED':
					ws[str(cellname)].font = ft1
			wb.save(file_name)

	# 将测试结果写入到Excel
	def resultAddExcel(self, file_name, sheet_name, cell_row, cell_column, count, *value):
		with closing(load_workbook(filename=file_name)) as wb:
			sheets = wb.sheetnames
			if sheet_name in sheets:
				ws = wb[str(sheet_name)]
			else:
				ws = wb.create_sheet(str(sheet_name))
			ft1 = Font(name="微软雅黑", color=colors.RED, size=10)
			ft2 = Font(name="微软雅黑", size=10, bold=True)
			ft3 = Font(name="微软雅黑 Light", size=9)
			if value[int(count)] == 'FAIL':
				ws.cell(row=int(cell_row), column=int(cell_column), value=value[int(count)]).font = ft1
			elif value[int(count)] == 'PASS':
				ws.cell(row=int(cell_row), column=int(cell_column), value=value[int(count)]).font = ft2
			else:
				ws.cell(row=int(cell_row), column=int(cell_column), value=value[int(count)]).font = ft3
			wb.save(file_name)

	# 将字典变量内容写入到Excel
	def dictAddTitleExcel(self, file_name, sheet_name, cell_row, dict):
		with closing(load_workbook(filename=file_name)) as wb:
			sheets = wb.sheetnames
			if sheet_name in sheets:
				ws = wb[str(sheet_name)]
			else:
				ws = wb.create_sheet(str(sheet_name))
			ft = Font(name="微软雅黑", size=10, bold=False)
			for key in dict.keys():
				val = dict[key]
				ws.cell(row=int(cell_row), column=int(key), value=str(val)).font = ft
			wb.save(file_name)

	# 读取Excel
	def read_excel_no_null(file_name, sheet_name, row_count):
		with closing(load_workbook(filename=file_name, data_only=True)) as wb:
			ws = wb[str(sheet_name)]
			colValueList = []
			num = int(row_count)
			for cell in list(ws.columns)[num]:
				if cell.value != None:
					values = cell.value
					colValueList.append(values)
				else:
					pass
			return colValueList

	# 功能：将一字典写入到csv文件中
	# 输入：文件名称，数据字典
	# def create_dict_csv(self, file_name="", data_dict={}):
	# 	with open(file_name, "wb") as csv_file:
	# 		csv_writer = csv.writer(csv_file)
	# 		for key in data_dict:
	# 			csv_writer.writerow([key, data_dict[key]])
	# 		csv_file.close()

	# def dict_to_csv(self, filename, dict):
	# 	csv_file = open(filename, 'w')
	# 	csv_file.write(codecs.BOM_UTF8)
	# 	writer = csv.writer(csv_file)
	# 	for key in dict:
	# 		writer.writerow([key, dict[key]])
	# 	csv_file.close()

	def list_dict_to_csv(self, filename, list_dict, row_name):
		def get_sorted_values(row):
			sorted_values = []
			keys = row.keys()
			keys.sort()
			for key in keys:
				sorted_values.append(row[key])
			return sorted_values

		csv_file = open(filename, 'ab')

		csv_file.write(codecs.BOM_UTF8)

		writer = csv.writer(csv_file)

		sorted_values = get_sorted_values(row_name)
		writer.writerow(sorted_values)

		for row in list_dict:
			sorted_values = get_sorted_values(row)
			writer.writerow(sorted_values)
		csv_file.close()

	def dict_to_csv_with_title(self, filename, dict, row_name):
		def get_sorted_values(row):
			sorted_values = []
			keys = row.keys()
			keys.sort()
			for key in keys:
				sorted_values.append(row[key])
			return sorted_values

		csv_file = open(filename, 'wb')

		csv_file.write(codecs.BOM_UTF8)

		writer = csv.writer(csv_file)

		sorted_values = get_sorted_values(row_name)
		writer.writerow(sorted_values)

		sorted_values1 = get_sorted_values(dict)
		writer.writerow(sorted_values1)

		csv_file.close()

	def dict_to_csv_with_title_batch(self, filename, dict_list, row_name):
		def get_sorted_values(row):
			sorted_values = []
			keys = row.keys()
			keys.sort()
			for key in keys:
				sorted_values.append(row[key])
			return sorted_values

		csv_file = open(filename, 'wb')

		csv_file.write(codecs.BOM_UTF8)

		writer = csv.writer(csv_file)
		sorted_values = get_sorted_values(row_name)
		writer.writerow(sorted_values)

		for i in range(len(dict_list)):
			sorted_values1 = get_sorted_values(dict_list[i])
			writer.writerow(sorted_values1)

		csv_file.close()

	def dict_to_csv_no_title(self, filename, dict):
		def get_sorted_values(row):
			sorted_values = []
			keys = row.keys()
			keys.sort()
			for key in keys:
				sorted_values.append(row[key])
			return sorted_values

		csv_file = open(filename, 'ab')

		csv_file.write(codecs.BOM_UTF8)

		writer = csv.writer(csv_file)

		sorted_values1 = get_sorted_values(dict)
		writer.writerow(sorted_values1)

		csv_file.close()

	def dict_to_csv_no_title_batch(self, filename, dict_list):
		def get_sorted_values(row):
			sorted_values = []
			keys = row.keys()
			keys.sort()
			for key in keys:
				sorted_values.append(row[key])
			return sorted_values

		csv_file = open(filename, 'ab')

		csv_file.write(codecs.BOM_UTF8)

		writer = csv.writer(csv_file)

		for i in range(len(dict_list)):
			sorted_values1 = get_sorted_values(dict_list[i])
			writer.writerow(sorted_values1)

		csv_file.close()


	def read_dict_csv(self, fileName=""):
		with open(fileName, 'rb') as csvfile:
			reader = csv.DictReader(csvfile)
			rows = [row for row in reader]
		return rows

	# 将字典变量内容写入到Excel
	def dict_write_excel_by_dict_key(self, file_name, sheet_name, cell_row, w_dict):
		with closing(load_workbook(filename=file_name)) as wb:

			dict = {'9': w_dict['实际返回码'], '11':w_dict['实际返回信息'], '12': w_dict['appId']}
			sheets = wb.sheetnames
			if sheet_name in sheets:
				ws = wb[str(sheet_name)]
			else:
				ws = wb.create_sheet(str(sheet_name))
			ft = Font(name="微软雅黑 Light", size=10, bold=False)
			for key in dict.keys():
				val = dict[key]
				ws.cell(row=int(cell_row), column=int(key), value=str(val)).font = ft
			wb.save(file_name)

	# 将字典变量内容写入到Excel
	def dictList_write_excel_by_dict_key_batch(self, file_name, sheet_name, cell_row_start, dict_list):
		with closing(load_workbook(filename=file_name)) as wb:

			sheets = wb.sheetnames
			if sheet_name in sheets:
				ws = wb[str(sheet_name)]
			else:
				ws = wb.create_sheet(str(sheet_name))
			ft = Font(name="微软雅黑 Light", size=10, bold=False)
			for i in range(len(dict_list)):
				dict = {'9': dict_list[i]['实际返回码'], '11': dict_list[i]['实际返回信息'],  '12': dict_list[i]['appId'], '13': dict_list[i]['SEND_DATA']}
				for key in dict.keys():
					val = dict[key]
					ws.cell(row=int(int(cell_row_start) + i), column=int(key), value=str(val)).font = ft
			wb.save(file_name)

	def dictList_write_excel_by_dict_key_batch_bak(self, file_name, sheet_name, cell_row_start, dict_list):
		with closing(load_workbook(filename=file_name)) as wb:

			sheets = wb.sheetnames
			if sheet_name in sheets:
				ws = wb[str(sheet_name)]
			else:
				ws = wb.create_sheet(str(sheet_name))
			ft = Font(name="微软雅黑 Light", size=10, bold=False)
			for i in range(len(dict_list)):
				dict = {'9': dict_list[i]['实际返回码'], '11': dict_list[i]['实际返回信息'], '12': dict_list[i]['appId'], '13': dict_list[i]['SEND_DATA']}
				for key in dict.keys():
					val = dict[key]
					ws.cell(row=int(int(cell_row_start) + i), column=int(key), value=str(val)).font = ft
			wb.save(file_name)

	def dictList_write_appId_excel_by_dict_key_batch(self, file_name, sheet_name, cell_row_start, dict_list):
		with closing(load_workbook(filename=file_name)) as wb:

			sheets = wb.sheetnames
			if sheet_name in sheets:
				ws = wb[str(sheet_name)]
			else:
				ws = wb.create_sheet(str(sheet_name))
			ft = Font(name="微软雅黑 Light", size=10, bold=False)
			for i in range(len(dict_list)):
				dict = {'1': dict_list[i]['appId']}
				for key in dict.keys():
					val = dict[key]
					ws.cell(row=int(int(cell_row_start) + i), column=int(key), value=str(val)).font = ft
			wb.save(file_name)

	def List_write_paraName_excel_by_dict_key_batch(self, file_name, sheet_name, cell_row_start, cols, list):
		with closing(load_workbook(filename=file_name)) as wb:

			sheets = wb.sheetnames
			if sheet_name in sheets:
				ws = wb[str(sheet_name)]
			else:
				ws = wb.create_sheet(str(sheet_name))
			ft = Font(name="微软雅黑 Light", size=10, bold=False)
			for i in range(len(list)):
				ws.cell(row=int(int(cell_row_start) + i), column=int(int(cols) + 1), value=str(list[i])).font = ft
			wb.save(file_name)



	def dictAddFormulaToExcel(self, file_name, sheet_name, cell_row, dict):
		with closing(load_workbook(filename=file_name)) as wb:
			sheets = wb.sheetnames
			if sheet_name in sheets:
				ws = wb[str(sheet_name)]
			else:
				ws = wb.create_sheet(str(sheet_name))
			ft = Font(name="微软雅黑", size=12, bold=True)
			for key in dict.keys():
				val = dict[key]
				formula = '''=IF(INDIRECT(ADDRESS(int(''' + str(
					cell_row) + '''),COLUMN()-1))=INDIRECT(ADDRESS(int(''' + str(
					cell_row) + '''),COLUMN()-2)),"Pass","Fail")'''
				ws.cell(row=int(cell_row), column=int(val), value=formula).font = ft
			wb.save(file_name)


	# 按列分段读取Excel
	def colname_read_excel_return_list(self, file_name, sheet_name, col_name, start, end):
		with closing(load_workbook(filename=file_name, data_only=True)) as wb:
#			ws = wb.get_sheet_by_name(str(sheet_name))
			ws = wb[str(sheet_name)]

			if int(start) == int(end):
				cellname = str(col_name) + str(start)
				cellValue = ws[str(cellname)].value
				cellValueList = [str(cellValue)]
			else:
				cellValueList = []
				for i in range(int(start), int(end)+1):
					cellname = str(col_name) + str(i)
					cellValue = ws[str(cellname)].value
					cellValueList.append(str(cellValue))
			return cellValueList

#For test
if __name__ == '__main__':
	config_excel = r'E:\自动化\05_BUS自动化工具开发\02_接口自动化测试框架\BusAutoTest_Version4\Template\Product_Para.xlsx'
	config_excel = unicode(config_excel, "utf-8")
	para_config_sheet ='parameter'
	excel = ReadWriteExcel()
	excel.get_cell_index(config_excel, para_config_sheet, 1)
